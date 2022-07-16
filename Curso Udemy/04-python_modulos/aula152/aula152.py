import openpyxl
from random import uniform

'''
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']
'''

#print(planilha1['b4'].value)  # selecionando o item da coluna B linha 4

#for linha in planilha1['a1:c2']:  # percorrendo do item A1 até o C2
#    for coluna in linha:
#        print(coluna.value)

#for linha in planilha1:  # percorrendo a planilha toda
#    for coluna in linha:
#        print(coluna.value)

# alterando um valor
#$planilha1['b3'].value = 2200
#pedidos.save('nova_planilha.xlsx')  # salvando a alteração em uma nova planilha

'''
for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    
    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco
'''

planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    
    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(5, 16):
    planilha2.cell(linha, 1).value = f'Andre {linha} {round(uniform(10, 100)), 2}'
    planilha2.cell(linha, 1).value = f'Joaozinho {linha} {round(uniform(10, 100)), 2}'
    planilha2.cell(linha, 1).value = f'Maria {linha} {round(uniform(10, 100)), 2}'

planilha.save('planilha_gerada.xlsx')
